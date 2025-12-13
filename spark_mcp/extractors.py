"""Text extraction utilities for email attachments."""

import base64
from pathlib import Path
from typing import Tuple


def extract_text(file_path: str, mime_type: str) -> Tuple[str, str]:
    """Extract text from a file.

    Args:
        file_path: Path to the file
        mime_type: MIME type of the file

    Returns:
        Tuple of (content, content_type) where content_type is 'extracted_text' or 'base64'
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Attachment not found: {file_path}")

    # PDF extraction
    if mime_type == "application/pdf":
        try:
            return extract_pdf(file_path), "extracted_text"
        except Exception as e:
            return f"[PDF extraction failed: {str(e)}]", "error"

    # Word documents
    if mime_type in [
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ]:
        try:
            return extract_docx(file_path), "extracted_text"
        except Exception as e:
            return f"[Word document extraction failed: {str(e)}]", "error"

    # Excel spreadsheets
    if mime_type in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]:
        try:
            return extract_xlsx(file_path), "extracted_text"
        except Exception as e:
            return f"[Excel extraction failed: {str(e)}]", "error"

    # Plain text files
    if mime_type.startswith("text/"):
        try:
            return path.read_text(errors="replace"), "extracted_text"
        except Exception as e:
            return f"[Text file read failed: {str(e)}]", "error"

    # Calendar files (iCal)
    if mime_type in ["text/calendar", "application/ics"]:
        try:
            return path.read_text(errors="replace"), "extracted_text"
        except Exception as e:
            return f"[Calendar file read failed: {str(e)}]", "error"

    # Binary files - return base64
    try:
        return base64.b64encode(path.read_bytes()).decode(), "base64"
    except Exception as e:
        return f"[Failed to read file: {str(e)}]", "error"


def extract_pdf(file_path: str) -> str:
    """Extract text from a PDF file.

    Args:
        file_path: Path to the PDF file

    Returns:
        Extracted text content
    """
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    pages_text = []

    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text:
            pages_text.append(f"--- Page {i + 1} ---\n{text}")

    if not pages_text:
        return "[No extractable text found in PDF]"

    return "\n\n".join(pages_text)


def extract_docx(file_path: str) -> str:
    """Extract text from a Word document.

    Args:
        file_path: Path to the .docx file

    Returns:
        Extracted text content
    """
    from docx import Document

    doc = Document(file_path)
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]

    if not paragraphs:
        return "[No extractable text found in document]"

    return "\n\n".join(paragraphs)


def extract_xlsx(file_path: str) -> str:
    """Extract text from an Excel spreadsheet.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        Extracted text content (tab-separated values per sheet)
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets_text = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            # Filter None values and convert to strings
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(v.strip() for v in row_values):
                rows.append("\t".join(row_values))

        if rows:
            sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    wb.close()

    if not sheets_text:
        return "[No extractable data found in spreadsheet]"

    return "\n\n".join(sheets_text)
